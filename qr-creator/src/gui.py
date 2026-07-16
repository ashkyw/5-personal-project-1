import qrcode
import openpyxl
import customtkinter as ctk
from tkinterdnd2 import DND_FILES, TkinterDnD

class TopLevelWindow(ctk.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("400x300")

        self.label = ctk.CTkLabel(self, text="TopLevelWindow")
        self.label.pack(padx=20, pady=20)

class LabelText(ctk.CTkFrame):
    def __init__(self, master, values):
        super().__init__(master)

        self.grid_columnconfigure(0, weight=10)
        self.values: list[str] = values

        for i, value in enumerate(self.values):
            self.label = ctk.CTkLabel(self, text=value, fg_color="transparent", justify="center")
            self.label.grid(row=i+1, column=0, padx=10, pady=(10, 0), sticky="ew")

class DragAndDropFrame(ctk.CTkFrame):
    def __init__(self, master, title, on_files=None, valid_extensions=None):
        super().__init__(master)
        self.title: str = title
        self.on_files: list[str] = on_files
        self.valid_extensions: tuple[str] = valid_extensions

        self.grid_columnconfigure(0, weight=10)

        self.title = ctk.CTkLabel(self, text=self.title, font=ctk.CTkFont(size=16, weight="bold"))
        self.title.grid(row=0, column=0, padx=10, pady=(10,0), sticky="ew")

        self.drop_target_register(DND_FILES)
        self.dnd_bind("<<Drop>>", self.on_drop)

    # ID and filter unwanted captured data
    def on_drop(self, event: list[str]) -> None:
        files = self.tk.splitlist(event.data)
        valid = self.validate_files(files, self.valid_extensions)

        if self.on_files:
            self.on_files(valid)

    # Validate files before displaying
    def validate_files(self, files: list[str], valid_extensions: tuple[str]):
        valid_files: list[str] = []
        for file in files:
            if file.endswith(self.valid_extensions):
                if file not in valid_files:
                    valid_files.append(file)

        return valid_files

class ButtonFrame(ctk.CTkFrame):
    # Takes text and commands as lists
    def __init__(self, master, title, text=None, commands=None):
        super().__init__(master)
        self.grid_columnconfigure(0, weight=1)
        self.title: str = title
        self.button_text: str = text
        self.buttons: list[str] = []
        self.commands: list[str] = commands

        self.title = ctk.CTkLabel(self, text=self.title, fg_color="gray30", corner_radius=6)
        self.title.grid(row=0, column=0, padx=10, pady=(10, 0), sticky="ew")
        if self.button_text is not None and self.commands is not None:
            for i, text in enumerate(self.button_text):
                button = ctk.CTkButton(self, text=text, command=self.commands[i])
                button.grid(row=i+1, column=0, padx=10, pady=(10, 0), sticky="ew")

class CheckBoxFrame(ctk.CTkFrame):
    def __init__(self, master, title, values):
        super().__init__(master)
        self.grid_columnconfigure(0, weight=1)
        self.values: list[str] = values
        self.title: str = title
        self.checkboxes: list[str, int] = []

        self.title = ctk.CTkLabel(self, text=self.title, fg_color="gray30", corner_radius=6)
        self.title.grid(row=0, column=0, padx=10, pady=(10, 0), sticky="ew")

        for i, value in enumerate(self.values):
            checkbox = ctk.CTkCheckBox(self, text=value)
            checkbox.grid(row=i+1, column=0, padx=10, pady=(10, 0), sticky="w")
            self.checkboxes.append(checkbox)

    def get(self) -> list[str, int]:
        checked_checkboxes: list[str, int] = []
        for checkbox in self.checkboxes:
            if checkbox.get() == 1:
                checked_checkboxes.append(checkbox.cget("text"))
        return checked_checkboxes

class RadiobuttonFrame(ctk.CTkFrame):
    def __init__(self, master, title, values):
        super().__init__(master)
        self.grid_columnconfigure(0, weight=1)
        self.values: list[str] = values
        self.title: str = title
        self.radiobuttons: list[str] = []
        self.variable = ctk.StringVar(value="")

        self.title = ctk.CTkLabel(self, text=self.title, fg_color="gray30", corner_radius=6)
        self.title.grid(row=0, column=0, padx=10, pady=(10, 0), sticky="ew")

        for i, value in enumerate(self.values):
            radiobutton = ctk.CTkRadioButton(self, text=value, value=value, variable=self.variable)
            radiobutton.grid(row=i+1, column=0, padx=10, pady=(10, 0), sticky="w")
            self.radiobuttons.append(radiobutton)

    def get(self) -> str:
        return self.variable.get()

    def set(self, value) -> None:
        self.variable.set(value)

class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        # Set App window settings
        self.title("App title goes here")
        self.geometry("800x640")
        self.grid_columnconfigure((0, 1), weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.configure(bg="gray30")

        self.files: list[str] = []

        self.toplevel_window = None

        self.button = ctk.CTkButton(self, text="New window", command=self.create_toplevel_window)
        self.button.grid(row=3, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

        # Set ButtonFrame buttons & settings
        self.button_frame = ButtonFrame(
            self, "Functions", text=["New Window", "Print Success"],
            commands=[self.create_toplevel_window, self.test_adding_commands_to_buttons]
        )
        self.button_frame.grid(row=0, column=0, padx=10, pady=(10,0), sticky="nsew")

        # Set DragAndDropFrame labels & settings
        self.drop_frame = DragAndDropFrame(
            self, "Drag files here...", on_files=self.files_dropped,
            # uncomment for testing purposes valid_extensions=(".pdf", ".png")
            valid_extensions=(".xlsx", ".xlsm", ".csv", ".xlsb")
        )
        self.drop_frame.grid(row=0, column=1, padx=(0,10), pady=(10,0), sticky="nsew", columnspan=2)
        self.drop_frame.configure(fg_color="transparent")

    # Hands files off to other functions
    def files_dropped(self, files: list[str]) -> None:
        self.files: list[str] = files

        # Display and update dropped files
        # If frame exists, clear it
        if hasattr(self, "drop_text_frame"):
            self.drop_text_frame.destroy()
            self.remove_frame.destroy()

        # Set child labels & values
        self.drop_text_frame = LabelText(self.drop_frame, self.files)
        self.drop_text_frame.grid(row=1, column=0, padx=10, pady=(10,0), sticky="ew")

        #self.create_remove_frame(files)

    # Creates button frame to remove file from files list
    def create_remove_frame(self, files: list[str]) -> None:
        # If you go with this function, you need to update remove_file_from_list to pop items from the list, then update the buttons with the new list indices.
        # Really, it's a lot of work for an algorithm that personal developmental-wise would be fantastic, but in reality is unnecessary when users can just re-drag the files into the window to reset
        self.remove_frame = ButtonFrame(self.drop_frame, "Remove Files")
        self.remove_frame.grid(row=1, column=1, padx=10, pady=(10,0), sticky="ew")
        if self.files != None:
            for i, file in enumerate(self.files):
                self.button = ctk.CTkButton(self.remove_frame, text="Remove", command=self.test_adding_commands_to_buttons)
                self.button.grid(row=i+1, column=1, padx=10, pady=(10,0), sticky="ew")

    # Button functions
    def create_toplevel_window(self) -> None:

        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = TopLevelWindow(self)
        else:
            self.toplevel_window.focus()

    def test_adding_commands_to_buttons(self) -> str:
        print("Every thing works")

    def remove_file_from_list(self, files: list [str]) -> None:
        # See note in create_remove_frame for more details
        #self.files.pop()
        PASS

    def create_qrcodes(files: list[str]) -> None:
        if files == []:
            # Setup some sort of error system. Easiest is likely a new top level window saying the list is empty
            PASS
        # Below will work for single files, but need to come up with a way to successfully iterate over multiple files at once
        # to ensure all files are correctly handled.
        for file in files:
            # set active workbook
            workbook = openpyxl.load_workbook(file)
            # set active worksheet
            ws = workbook.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row != (None,):
                    cleaned_row = "".join(map(str, row))
                    img = qrcode.make(cleaned_row)
                    img_column = ws.cell(row=i+1, column=2)
                    img_column.value = img
                    #img.save(f"{i}.png")

        '''
        for row in ws.iter_rows(values_only=True):
            if row != (None,):
                cleaned_row = "".join(map(str, row))
                img = qrcode.make(cleaned_row)
                img.save(f"{i}.png")
                i += 1
        '''

    def button_callback(self) -> None:
        print("checked checkboxes", self.checkbox_frame.get())
        print("radiobutton_frame", self.radiobutton_frame.get())
